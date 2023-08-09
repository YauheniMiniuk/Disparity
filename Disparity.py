import numpy as np
import pandas as pd
from Levenshtein import distance as levenshtein_distance
from collections import Counter
from keras.utils.data_utils import pad_sequences
from sklearn.metrics.pairwise import cosine_similarity
from pyjarowinkler import distance as jaro_winkler_distance
import re
import string as s
import random
import torch
from transformers import AutoTokenizer, AutoModel
from gensim.models import FastText
import nltk
from sklearn.feature_extraction.text import CountVectorizer
import openpyxl


RB_PRICES_PATH = 'Average_prices-06-2023.xls'
RF_PRICES_PATH = 'data.xls'

# Класс Product, который имеет свойства Имя, Цены, Доля, Родительский класс
class Product:
    def __init__(self, name):
        self.name = name
        self.weight = None
        self.pricesBYN = dict()
        self.pricesRUB = dict()
        self.subgroup = None
        self.group = None

    def SetGroup(self, group):
        self.group = group

    def GetGroup(self):
        return self.group

    def SetSubgroup(self, subgroup):
        self.subgroup = subgroup

    def GetSubgroup(self):
        return self.subgroup

    def SetWeight(self, weight):
        self.weight = weight

    def SetPriceBYN(self, state, price):
        self.pricesBYN[state] = price

    def SetPriceRUB(self, state, price):
        self.pricesRUB[state] = price

# Класс Subgroup, который имеет Имя, список объектов Product, метод, который возвращает список объектов Product


class Subgroup:
    def __init__(self, name):
        self.name = name
        self.products = []  # список объектов Product
        self.group = None

    # Метод, который возвращает список объектов Product
    def GetProducts(self):
        return self.products

    def GetGroup(self):
        return self.group

    def SetGroup(self, group):
        self.group = group

    # Функция Append, которая добавляет объект Product в список products
    def Append(self, product):
        if isinstance(product, Product):  # проверяем, что аргумент является объектом Product
            self.products.append(product)  # добавляем его в список products
        else:
            print("Invalid argument")  # иначе выводим сообщение об ошибке

# Класс Group, который имеет Имя, список объектов Product, список объектов Subgroup,
# имеет метод, который возвращает список объектов Product, который содержит сам и который содержится в объектах Subgroup


class Group:
    def __init__(self, name):
        self.name = name
        self.products = []  # список объектов Product
        self.subgroups = []  # список объектов Subgroup

    # Метод, который возвращает список объектов Product, который содержит сам и который содержится в объектах Subgroup
    def GetProducts(self):
        result = []  # пустой список для результата
        result.extend(self.products)  # добавляем в него свои продукты
        for subgroup in self.subgroups:  # проходим по подгруппам
            # добавляем в него продукты из подгрупп
            result.extend(subgroup.GetProducts())
        return result

    def GetSubgroups(self):
        return self.subgroups

    # Функция Append, которая добавляет объект Product или Subgroup в соответствующий список
    def Append(self, item):
        if isinstance(item, Product):  # проверяем, что аргумент является объектом Product
            self.products.append(item)  # добавляем его в список products
        elif isinstance(item, Subgroup):  # проверяем, что аргумент является объектом Subgroup
            self.subgroups.append(item)  # добавляем его в список subgroups
        else:
            print("Invalid argument")  # иначе выводим сообщение об ошибке

# Класс Category, который имеет Имя и метод получения списка всех объектов Product у всех дочерних объектов


class Category:
    def __init__(self, name):
        self.name = name
        self.groups = []  # список объектов Group

    # Метод получения списка всех объектов Product у всех дочерних объектов
    def GetProducts(self):
        result = []  # пустой список для результата
        for group in self.groups:  # проходим по группам
            # добавляем в него продукты из групп и подгрупп
            result.extend(group.GetProducts())
        return result

    def GetGroups(self):
        return self.groups

    # Функция Append, которая добавляет объект Group в список groups
    def Append(self, group):
        if isinstance(group, Group):  # проверяем, что аргумент является объектом Group
            self.groups.append(group)  # добавляем его в список groups
        else:
            print("Invalid argument")  # иначе выводим сообщение об ошибке

# определяем функцию для получения части речи слова


def get_pos(words):
    # получаем кортеж (слово, часть речи)
    tags = nltk.pos_tag(words, lang='rus')
    for tag in tags:
        # преобразуем часть речи в удобный формат
        if tag[1].lower() == 's':
            yield (tag[0], 'noun')
        elif tag[1].lower() == 'v':
            yield (tag[0], 'noun')
        elif tag[1].lower() == 'a=f':
            yield (tag[0], 'adj')
        elif tag[1].lower() == 'a=n':
            yield (tag[0], 'adj')
        elif tag[1].lower() == 'r':
            yield (tag[0], 'adv')
        else:
            yield (tag[0], None)

# определяем функцию для получения веса слова в зависимости от его части речи


def get_weight(pos):
    # задаем словарь весов для разных частей речи
    weights = {'noun': 10, 'adj': 7, 'verb': 5, 'adv': 2, None: 0}
    # если часть речи есть в словаре, возвращаем соответствующий вес
    if pos in weights:
        return weights[pos]
    # иначе возвращаем 1
    else:
        return 1

# определяем функцию для получения списка слов с весами из строки


def get_words_with_weights(string):
    # разбиваем строку на слова
    words = nltk.word_tokenize(string)
    # создаем пустой список для хранения слов с весами
    words_with_weights = []
    poss = get_pos(words)
    # для каждого слова в строке
    for pos in poss:
        # получаем его вес
        weight = get_weight(pos[1])
        # добавляем слово с весом в список
        words_with_weights.append((pos[0], weight))
    # возвращаем список слов с весами
    return words_with_weights

# определяем функцию для получения косинусного сходства между двумя строками по новому алгоритму


def get_cosine_similarity_new(string1, string2):
    # получаем списки слов с весами из строк
    words_with_weights1 = get_words_with_weights(string1)
    words_with_weights2 = get_words_with_weights(string2)

    # создаем пустые списки для хранения слов и весов отдельно
    words1 = []
    weights1 = []
    words2 = []
    weights2 = []

    # для каждого слова с весом из первого списка
    for word, weight in words_with_weights1:
        # добавляем слово в список слов первой строки
        words1.append(word)
        # добавляем вес в список весов первой строки
        weights1.append(weight)

     # аналогично для каждого слова с весом из второго списка
    for word, weight in words_with_weights2:
        words2.append(word)
        weights2.append(weight)

    # создаем объект CountVectorizer для преобразования слов в числовые признаки
    # устанавливаем параметр lowercase равным False, чтобы сохранить регистр слов
    vectorizer = CountVectorizer()

    # объединяем списки слов из обеих строк в один список
    words = words1 + words2

    # преобразуем список слов в матрицу признаков X
    X = vectorizer.fit_transform(words)

    # получаем список имен признаков (уникальных слов) из объекта vectorizer
    features = vectorizer.get_feature_names_out()

    # создаем пустые списки для хранения признаков с весами для обеих строк
    features_with_weights1 = []
    features_with_weights2 = []

    # для каждого признака (слова) в списке признаков
    for feature in features:
        # если признак есть в списке слов первой строки
        if feature in words1:
            # получаем индекс признака в списке слов первой строки
            index = words1.index(feature)
            # получаем вес признака из списка весов первой строки по индексу
            weight = weights1[index]
        # иначе
        else:
            # задаем вес признака равным 0
            weight = 0
        # добавляем признак с весом в список признаков с весами для первой строки
        features_with_weights1.append((feature, weight))

        # аналогично для второй строки
        if feature in words2:
            index = words2.index(feature)
            weight = weights2[index]
        else:
            weight = 0
        features_with_weights2.append((feature, weight))

    # создаем пустые списки для хранения сопоставленных частей между строками и их суммарных весов
    matched_parts = []
    matched_weights = []

    # для каждого признака с весом из первого списка
    for feature, weight in features_with_weights1:
        # если признак не равен 0 (то есть присутствует в первой строке)
        if weight != 0:
            # если признак есть во втором списке и его вес не равен 0 (то есть присутствует во второй строке)
            if feature in dict(features_with_weights2) and dict(features_with_weights2)[feature] != 0:
                # добавляем признак в список сопоставленных частей между строками
                matched_parts.append(feature)
                # добавляем сумму весов признака из обеих строк в список сопоставленных весов
                matched_weight = weight + dict(features_with_weights2)[feature]
                matched_weights.append(matched_weight)

    # вычисляем общий результат косинусного сходства по сопоставленным частям и их весам
    cosine_similarity_result = sum(matched_weights)

    # вычисляем максимальный результат косинусного сходства, который может быть достигнут при полном совпадении строк
    cosine_similarity_max = sum(weights1) + sum(weights2)

    if cosine_similarity_result == 0 or cosine_similarity_max == 0:
        return 0
    # вычисляем соотношение между результатом и максимумом косинусного сходства
    cosine_similarity_ratio = cosine_similarity_result / cosine_similarity_max

    # округляем соотношение до двух знаков после запятой
    cosine_similarity_ratio = round(cosine_similarity_ratio, 2)

    # возвращаем соотношение как результат функции
    return cosine_similarity_ratio


df = pd.read_excel(RB_PRICES_PATH, header=6)
df = df.rename(columns={df.columns[0]: 'Название товара'})

prodPricesStartIndex = df[df['Название товара'].str.lower(
) == 'продовольственные товары'].index[0]
prodPricesEndIndex = df[df['Название товара'].str.lower(
) == 'непродовольственные товары'].index[0]
dfProdPrices = df[prodPricesStartIndex+1:prodPricesEndIndex]
# Удаляет все строки, где хоть в одной ячейке есть nan. Заменить на удаление только тех строк, где вообще все ячейки пустые
dfProdPrices = dfProdPrices.dropna()


noProdPricesStartIndex = df[df['Название товара'].str.lower(
) == 'непродовольственные товары'].index[0]
dfNoProdPrices = df[noProdPricesStartIndex+1:]
# Удаляет все строки, где хоть в одной ячейке есть nan. Заменить на удаление только тех строк, где вообще все ячейки пустые
dfNoProdPrices = dfNoProdPrices.dropna()


df2 = pd.read_excel("! Копия Веса.xlsx", sheet_name='РБ_Струкутура', header=8)
dfProdWeights = df2.iloc[:df2[df2['ПРОДОВОЛЬСТВЕННЫЕ ТОВАРЫ (без табачных изделий)']
                              == 'НЕПРОДОВОЛЬСТВЕННЫЕ ТОВАРЫ'].index[0]]
df2 = pd.read_excel("! Копия Веса.xlsx", sheet_name='РБ_Струкутура', header=8)
# df = df.rename(columns={'ПРОДОВОЛЬСТВЕННЫЕ ТОВАРЫ (без табачных изделий)':'name'})
dfNoProdWeights = df2.iloc[df2[df2['ПРОДОВОЛЬСТВЕННЫЕ ТОВАРЫ (без табачных изделий)']
                               == 'НЕПРОДОВОЛЬСТВЕННЫЕ ТОВАРЫ'].index[0]+1:]


wb = openpyxl.load_workbook("! Копия Веса.xlsx")
ws = wb.get_sheet_by_name('РБ_Струкутура')


def FindIndexByText(ws, column, text):
    for i in range(9, ws.max_row + 1):
        # Получаем значение ячейки в первом столбце
        value = ws.cell(i, column).value
        # Сравниваем значение с искомым текстом
        if value != None and text.lower() in value.lower():
            # Выводим индекс строки
            return i


prodIndex = FindIndexByText(ws, 1, 'Продовольственные товары')
noProdIndex = FindIndexByText(ws, 1, 'Непродовольственные товары')
serviceIndex = FindIndexByText(ws, 1, 'Услуги')


prodCategoryRB = Category('Продовольственные товары')

lastGroup = None
lastSubgroup = None

for i in range(prodIndex+1, noProdIndex):
    cell = ws.cell(i, 1)
    weightCell = ws.cell(i, 2)
    value = cell.value
    weight = weightCell.value
    font = cell.font
    if weight:
        if font.bold and not font.italic and not font.underline:
            print('Group: ', value)
            group = Group(value)
            prodCategoryRB.Append(group)
            lastGroup = group
            lastSubgroup = None
        elif not font.bold and font.italic and font.underline:
            print('Subgroup', value)
            subgroup = Subgroup(value)
            group.Append(subgroup)
            lastSubgroup = subgroup
        elif not font.bold and not font.italic and not font.underline:
            print('Product:', value, '-', weight)
            product = Product(value)
            product.SetWeight(weight)
            if lastSubgroup != None:
                product.SetSubgroup(lastSubgroup)
                product.SetGroup(lastGroup)
                lastSubgroup.Append(product)
            else:
                product.SetGroup(lastGroup)
                lastGroup.Append(product)


noProdCategoryRB = Category('Непродовольственные товары')


lastGroup = None
lastSubgroup = None

for i in range(noProdIndex+1, serviceIndex):
    cell = ws.cell(i, 1)
    value = cell.value
    weightCell = ws.cell(i, 2)
    weight = weightCell.value
    font = cell.font

    if weight:
        if font.bold and not font.italic and not font.underline:
            print('Group: ', value)
            group = Group(value)
            noProdCategoryRB.Append(group)
            lastGroup = group
            lastSubgroup = None
        elif font.bold and font.italic:
            print('Subgroup', value)
            subgroup = Subgroup(value)
            group.Append(subgroup)
            lastSubgroup = subgroup
        elif not font.bold and not font.italic and not font.underline:
            print('Product:', value, '-', weight)
            product = Product(value)
            product.SetWeight(weight)
            if lastSubgroup != None:
                product.SetSubgroup(lastSubgroup)
                product.SetGroup(lastGroup)
                lastSubgroup.Append(product)
            else:
                product.SetGroup(lastGroup)
                lastGroup.Append(product)


def loadProdPricesRB(path='Average_prices-06-2023.xls', header=7):
    ''' Загрузить продовольственные товары РБ из файла с ценами
    '''
    df = pd.read_excel(path, header=header)
    prod = df.iloc[:df[df['ПРОДОВОЛЬСТВЕННЫЕ ТОВАРЫ']
                       == 'НЕПРОДОВОЛЬСТВЕННЫЕ ТОВАРЫ'].index[0]]
    prod = prod.dropna()
    prod = prod.rename(columns={prod.columns[0]: 'Название товара', prod.columns[1]: 'РБ', prod.columns[2]: 'Брестская', prod.columns[3]: 'Витебская',
                       prod.columns[4]: 'Гомельская', prod.columns[5]: 'Гродненская', prod.columns[6]: 'Минск', prod.columns[7]: 'Минская', prod.columns[8]: 'Могилевская'})
    return prod


def loadNotProdPricesRB(path='Average_prices-06-2023.xls', header=7):
    ''' Загрузить непродовольственные товары РБ из файла с ценами
    '''
    df = pd.read_excel(path, header=header)
    notprod = df.iloc[df[df['ПРОДОВОЛЬСТВЕННЫЕ ТОВАРЫ']
                         == 'НЕПРОДОВОЛЬСТВЕННЫЕ ТОВАРЫ'].index[0]+1:]
    notprod = notprod.dropna()
    notprod = notprod.rename(columns={notprod.columns[0]: 'Название товара', notprod.columns[1]: 'РБ', notprod.columns[2]: 'Брестская', notprod.columns[3]: 'Витебская',
                             notprod.columns[4]: 'Гомельская', notprod.columns[5]: 'Гродненская', notprod.columns[6]: 'Минск', notprod.columns[7]: 'Минская', notprod.columns[8]: 'Могилевская'})
    return notprod


def clean_string(string):
    '''Функция очистки
    Получает строку, переводит в нижний регистр, очищает лишних данных, пробелов, знаков препинания и т.д.
    '''
    string = str(string).lower()
    # Ищем и заменяем скобки и все, что в них, на пустую строку
    string = re.sub(r"\(.*\)", " ", string)

    # Удаляем знаки препинания
    string = string.translate(str.maketrans(' ', ' ', s.punctuation))

    # Удаляем единицы измерения
    # шаблон регулярного выражения
    pattern = r'(\d+)*\s+(кг|г|л|мл|шт|штук|м|м2|м3|см|таблеток|пакетиков)\b'
    # замена единиц измерения на пустую строку
    string = re.sub(pattern, " ", string)
    # Удаляем цифры
    string = re.sub(r'\d+', " ", string)
    # Удаляем лишние пробелы
    string = re.sub(r"\s+", " ", string)

    # Удаляем лишние пробелы в начале и конце строки
    string = string.strip()

    # Добавляем очищенную строку в список
    return string


def clean(df, column):
    '''Применяет функцию clean_string() для всего столбца и записывает новые строки в новый столбец с припиской _cleaned'''
    df[column+'_cleaned'] = df[column].apply(lambda x: clean_string(x))
    return df


prodPricesRB = loadProdPricesRB()
prodPricesRB = clean(prodPricesRB, prodPricesRB.columns[0])
notProdPricesRB = loadNotProdPricesRB()
notProdPricesRB = clean(notProdPricesRB, notProdPricesRB.columns[0])


prodRB = prodCategoryRB.GetProducts()
noProdRB = noProdCategoryRB.GetProducts()


def my_levenshtein_distance(str1, str2):
    '''Функция для получения разницы между строками методом Левенштейна (не тот мужик из американского пирога),
      при котором строка разбивается на слова и сортируется по алфавиту
    '''
    str1_words = str1.split()
    str2_words = str2.split()

    str1_words.sort()
    str2_words.sort()

    str1_sorted = ' '.join(str1_words)
    str2_sorted = ' '.join(str2_words)

    return levenshtein_distance(str1_sorted, str2_sorted)


def combine_cosine_levenshtein_jar_win(str1, str2, alpha=0.25, beta=0.25, gamma=0.5):
    '''Функция сопоставления двух строк str1 и str2. Возвращает значение от 0 до 1, насколько похожи строки.
    Функция использует комбинацию из методов косинусного сопоставления, Левенштейна и Джаро-Винклера\n
    alpha: вес для метода косинусного сопоставления\n
    beta: вес для метода Левенштейна
    gamma: вес для метода Джаро-Винклера'''
    # convert strings to vectors of character frequencies
    vec1 = list(Counter(str1).values())
    vec2 = list(Counter(str2).values())
    # pad the vectors with zeros to the same length
    vec1, vec2 = pad_sequences([vec1, vec2], padding='post')
    # compute cosine similarity between vectors
    cos_sim = cosine_similarity([vec1], [vec2])[0][0]
    # normalize cosine similarity to [0, 1] range
    cos_sim = (1 + cos_sim) / 2
    # compute levenshtein distance between strings
    lev_dist = my_levenshtein_distance(str1, str2)
    # normalize levenshtein distance to [0, 1] range
    lev_dist = 1 - lev_dist / max(len(str1), len(str2))
    # compute jaro-winkler distance between strings
    jar_win_dist = jaro_winkler_distance.get_jaro_distance(str1, str2)
    # combine the three methods by weighted sum
    return alpha * cos_sim + beta * lev_dist + gamma * jar_win_dist


def GetProductCount(name):
    count_str = re.search(
        r"(\d+((,+|\.+)*\d*)?\s*((мг)|(г)|(гр)|(кг)|(мл)|(л)))|(1/\d+)", name)
    if count_str == None:
        return 1
    try:
        count_str = count_str.group(0)
        count_str = count_str.replace(",", ".")
        if re.search("\s*вес[\s$]+", count_str) != None:
            return 1
        if re.search("1/\d+", count_str) != None:
            count = count_str.replace("1/", "")
            return float(count) / 1000
        if re.search("\d+((,+|\.+)*\d*)?\s*(мг)", count_str) != None:
            count = count_str.replace("мг", "").strip()
            return float(count) / 1000000
        if re.search("\d+((,+|\.+)*\d*)?\s*(кг)", count_str) != None:
            count = count_str.replace("кг", "").strip()
            return float(count)
        if re.search("\d+((,+|\.+)*\d*)?\s*(гр)", count_str) != None:
            count = count_str.replace("гр", "").strip()
            return float(count) / 1000
        if re.search("\d+((,+|\.+)*\d*)?\s*(г)", count_str) != None:
            count = count_str.replace("г", "").strip()
            return float(count) / 1000
        if re.search("\d+((,+|\.+)*\d*)?\s*(мл)", count_str) != None:
            count = count_str.replace("мл", "").strip()
            return float(count) / 1000
        if re.search("\d+((,+|\.+)*\d*)?\s*(л)", count_str) != None:
            count = count_str.replace("л", "").strip()
            return float(count)
    except Exception as e:
        print(e)
        return 1


def compareRB(list1, df2, colname2, colname2prices, perc, alpha, beta, gamma, noProd=False):
    list2 = df2[colname2].to_list()
    resultdf = pd.DataFrame()
    for i in range(len(list1)):
        name1 = list1[i].name
        if noProd:
            name1 = list1[i].name + ' ' + list1[i].parrent.name
        name1Clean = clean_string(name1)
        min_dist = 0
        for j in range(len(list2)):
            score = combine_cosine_levenshtein_jar_win(
                name1Clean, clean_string(list2[j]), alpha, beta, gamma)
            if score > min_dist:
                min_dist = score
                similar_index = j

        if min_dist >= perc:
            count = GetProductCount(list2[similar_index])
            state = colname2prices
            price = float(df2[colname2prices].iloc[similar_index]) / count
            list1[i].SetPriceBYN(state, price)
            resultdf = resultdf.append({'name1': name1,
                                        'name2': list2[similar_index],
                                        'score': min_dist,
                                        'weight': list1[i].weight,
                                        'priceBYN': list1[i].pricesBYN}, ignore_index=True)
        else:
            resultdf = resultdf.append(
                {'name1': list1[i].name}, ignore_index=True)
    return resultdf


def compareRF(list1, df2, colname2, colname2prices, perc, alpha, beta, gamma, noProd=False):
    list2 = df2[colname2].to_list()
    resultdf = pd.DataFrame()
    for i in range(len(list1)):
        name1 = list1[i].name
        # if noProd:
        #     name1 = list1[i].name + ' ' + list1[i].parrent.name
        name1Clean = clean_string(name1)
        min_dist = 0
        for j in range(len(list2)):
            # score = match_products(list1[i].name, list2[j], 0, 0, None, None, None, None)
            score = combine_cosine_levenshtein_jar_win(
                name1Clean, clean_string(list2[j]), alpha, beta, gamma)
            if score > min_dist:
                min_dist = score
                similar_index = j
        if min_dist >= perc:
            state = colname2prices
            count = GetProductCount(list2[similar_index])
            price = float(df2[colname2prices].iloc[similar_index]) / count
            list1[i].SetPriceRUB(state, price)
            resultdf = resultdf.append({'name1': name1,
                                        'name2': list2[similar_index],
                                        'score': min_dist,
                                        'weight': list1[i].weight,
                                        'priceRUB': list1[i].pricesRUB}, ignore_index=True)
        else:
            resultdf = resultdf.append(
                {'name1': list1[i].name}, ignore_index=True)
    return resultdf


comparedProdRB = compareRB(prodRB, prodPricesRB, prodPricesRB.columns[0], prodPricesRB.columns[-2],
                           0.75, alpha=0.3, beta=0.45, gamma=0.25)


comparednoProdRB = compareRB(noProdRB, notProdPricesRB, notProdPricesRB.columns[-1], notProdPricesRB.columns[-2],
                             0.75, alpha=0.45, beta=0.55, gamma=0)


comparedProdRB.to_excel('prod_rb.xlsx')
comparednoProdRB.to_excel('noprod_rb.xlsx')


# Загружаем продовольственные и непродовольственные товары РФ
rf = pd.read_excel('! Копия Веса.xlsx', sheet_name='РФ_структура', header=3)
rf = rf.drop(axis=1, columns=[rf.columns[2], rf.columns[3], rf.columns[4], rf.columns[5],
             rf.columns[6], rf.columns[7], rf.columns[8], rf.columns[9], rf.columns[10]])


prodIndexRF = rf.loc[rf.iloc[:, 0] == '01'].index[0]
noProdIndexRF = rf.loc[rf.iloc[:, 0] == '03'].index[0]
servicesIndexRF = rf.loc[rf.iloc[:, 0] == '04'].index[0]
noProdIndexRF2 = rf.loc[rf.iloc[:, 0] == '05'].index[0]
servicesIndexRF2 = rf.loc[rf.iloc[:, 0] == '06'].index[0]

prodCategoryRF = Category('Продовольственные товары')
noProdCategoryRF = Category('Непродовольственные товары')

lastGroup = None
lastSubgroup = None

for i in range(prodIndexRF+1, noProdIndexRF):
    cell = rf.iloc[i, 0]
    name = rf.iloc[i, 1]
    index = cell.split('.')
    if len(index) == 3:
        print('Group: ', name)
        group = Group(name)
        prodCategoryRF.Append(group)
        lastGroup = group
        lastSubgroup = None
    elif len(index) == 4:
        print('Subgroup', name)
        subgroup = Subgroup(name)
        group.Append(subgroup)
        lastSubgroup = subgroup
    elif len(index) == 5 or len(index) == 6:
        print('Product:', name, '-', weight)
        product = Product(name)
        product.SetWeight(weight)
        if lastSubgroup != None:
            product.SetSubgroup(lastSubgroup)
            product.SetGroup(lastGroup)
            lastSubgroup.Append(product)
        else:
            product.SetGroup(lastGroup)
            lastGroup.Append(product)

lastGroup = None
lastSubgroup = None

for i in range(noProdIndexRF+1, servicesIndexRF):
    cell = rf.iloc[i, 0]
    name = rf.iloc[i, 1]
    index = cell.split('.')

    if len(index) == 3:
        print('Group: ', name)
        group = Group(name)
        noProdCategoryRF.Append(group)
        lastGroup = group
        lastSubgroup = None
    elif len(index) == 4:
        print('Subgroup', name)
        subgroup = Subgroup(name)
        group.Append(subgroup)
        lastSubgroup = subgroup
    elif len(index) == 5 or len(index) == 6:
        print('Product:', name, '-', weight)
        product = Product(name)
        product.SetWeight(weight)
        if lastSubgroup != None:
            product.SetSubgroup(lastSubgroup)
            product.SetGroup(lastGroup)
            lastSubgroup.Append(product)
        else:
            product.SetGroup(lastGroup)
            lastGroup.Append(product)

lastGroup = None
lastSubgroup = None

for i in range(noProdIndexRF2+1, servicesIndexRF2):
    cell = rf.iloc[i, 0]
    name = rf.iloc[i, 1]
    index = cell.split('.')

    if len(index) == 3:
        print('Group: ', name)
        group = Group(name)
        noProdCategoryRF.Append(group)
        lastGroup = group
        lastSubgroup = None
    elif len(index) == 4:
        print('Subgroup', name)
        subgroup = Subgroup(name)
        group.Append(subgroup)
        lastSubgroup = subgroup
    elif len(index) == 5 or len(index) == 6:
        print('Product:', name, '-', weight)
        product = Product(name)
        product.SetWeight(weight)
        if lastSubgroup != None:
            product.SetSubgroup(lastSubgroup)
            product.SetGroup(lastGroup)
            lastSubgroup.Append(product)
        else:
            product.SetGroup(lastGroup)
            lastGroup.Append(product)


def GetDisparity(price1, price2, rate, weight1=1, weight2=1):
    '''
    price1 - цена товара в первой валюте\n
    price2 - цена товара во второй валюте\n
    rate - курс валюты\n
    weight1 - вес товара в первой корзине\n
    weight2 - вес товара во второй корзине
    '''

    return (price1 * weight1 / rate) / (price2 * weight2)


def random_weights():
    '''Возвращает случайные веса для трёх методов
    '''
    alpha = random.random()
    beta = random.random()
    gamma = random.random()
    total = alpha + beta + gamma
    alpha = round(alpha / total, 2)
    beta = round(beta / total, 2)
    gamma = round(gamma / total, 2)
    return alpha, beta, gamma


def compare_rb_rf(list1, list2, perc, rate, alpha=0.25, beta=0.5, gamma=0.25, noProd=False):
    '''Функция нахождения максимально похожих строк, используя цикл и случайные веса
    '''
    # Создаем пустой датафрейм
    resultdf = pd.DataFrame()
    # Получаем список областей из цен РБ и РФ
    states_rb = set()
    states_rf = set()
    for item in list1:
        if item.pricesBYN != None:
            states_rb.update(item.pricesBYN.keys())
    for item in list2:
        if item.pricesRUB != None:
            states_rf.update(item.pricesRUB.keys())
    states_rb = list(states_rb)
    states_rf = list(states_rf)
    # Используем цикл сопоставления
    for i in range(len(list1)):
        min_dist = 0
        max_score = 0
        name1 = clean_string(list1[i].name)
        if noProd and list1[i].subgroup != None:
            name1 = clean_string(list1[i].name + ' ' + list1[i].subgroup.name)
        for j in range(len(list2)):
            name2 = clean_string(list2[j].name)

            score = get_cosine_similarity_new(name1, name2)
            # score = compare_all_ngrams_and_return_max_similarity(name1, name2)
            # score = combine_cosine_levenshtein_jar_win(name1, name2, alpha, beta, gamma)
            # score = match_products(list1[i].name, list2[j].name, 0, 0, None, None, None, None)
            if score > min_dist:
                max_score = score
                min_dist = score
                similar_index = j
        # Если нашли похожий товар и у него есть цена в РБ
        if min_dist >= perc and list1[i].pricesBYN != None:
            # Создаем словарь для записи в датафрейм
            print('Товар РБ:', list1[i].name,
                  'Товар РФ:', list2[similar_index].name)
            row_dict = {
                'Товар РБ': list1[i].name, 'Доля РБ': list1[i].weight, 'Товар РФ': list2[similar_index].name}
            # Для каждой пары областей считаем диспаритет
            for state_rb in states_rb:
                price_rb = list1[i].pricesBYN.get(state_rb, None)
                row_dict[state_rb] = price_rb
                for state_rf in states_rf:
                    price_rf = list2[similar_index].pricesRUB.get(
                        state_rf, None)
                    if price_rf != None and price_rb != None:
                        list1[i].SetPriceRUB(state_rf, price_rf)
                        row_dict[state_rf] = price_rf
                        # if list1[i].weight != None and list2[similar_index].weight != None:
                        # Стоимость рубля временная. Добавить функцию получения автоматически
                        disparity = GetDisparity(price_rb, price_rf, rate)
                        row_dict['Диспаритет'] = disparity
                        # else:
                        #     row_dict['Диспаритет'] = ''
                    else:
                        row_dict[state_rf] = ''
                        row_dict['Диспаритет'] = ''
            # Записываем в итоговый датафрейм
            resultdf = resultdf.append(row_dict, ignore_index=True)
        # else:
        #     # Если не нашли похожий товар или у него нет цены в РБ, то заполняем пустыми значениями
        #     row_dict = {'Товар РБ': list1[i].name, 'Доля РБ': list1[i].weight, 'Товар РФ': ""}
        #     for state_rb in states_rb:
        #         row_dict[state_rb] = ''
        #         for state_rf in states_rf:
        #             row_dict[state_rf] = ''
        #             row_dict['Диспаритет'] = ''
        #     resultdf = resultdf.append(row_dict, ignore_index=True)
        # Если это не последний товар или если он относится к другой группе товаров, то добавляем строку с названием группы товаров
        if i < len(list1) - 1 and (list1[i+1].group != list1[i].group):
            group_row = {'Товар РБ': list1[i+1].group.name.upper()}
            resultdf = resultdf.append(group_row, ignore_index=True)
    return resultdf


rf = pd.read_excel(RF_PRICES_PATH, header=2)

rf = rf.rename(columns={'Unnamed: 0': 'name'})
rf = clean(rf, 'name')


prodRF = prodCategoryRF.GetProducts()
noProdRF = noProdCategoryRF.GetProducts()


comparedProdRF = compareRF(prodRF, rf, rf.columns[0], rf.columns[1],
                           0.7, alpha=0.3, beta=0.5, gamma=0.2)


comparedNoProdRF = compareRF(noProdRF, rf, rf.columns[0], rf.columns[1],
                             0.7, alpha=0.35, beta=0.55, gamma=0.1)


prod_rb_rf = compare_rb_rf(prodRB, prodRF, 0.1, 0.035618, 0.35, 0.6, 0.05)


noProd_rb_rf = compare_rb_rf(
    noProdRB, noProdRF, 0.1, 0.035618, 0.27, 0.68, 0.05, noProd=True)

result = pd.concat([prod_rb_rf, noProd_rb_rf])
result.to_excel('result.xlsx')
